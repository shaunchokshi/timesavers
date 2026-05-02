#!/usr/bin/env python3
"""Convert between XLSX and CSV / TSV / Markdown / JSON.

Read mode (xlsx → tabular):
  CSV / TSV outputs are per-sheet: one file per worksheet, written into the
  directory passed via --out. Multi-sheet workbooks use
  "{stem}__{SheetName}.{ext}"; single-sheet workbooks use "{stem}.{ext}".
  Markdown and JSON outputs are combined: one file with each sheet rendered
  as an H2 section (markdown) or an entry in a "sheets" array (json). Pass
  --out as a file path, or omit it to write to stdout.

Write mode (tabular → xlsx):
  Use --to xlsx with a csv/tsv/md input. csv/tsv produces a single-sheet
  workbook named after the file stem. md is parsed leniently: every
  pipe-table found becomes a sheet, named after the most recent `## Heading`
  above it (or Sheet1, Sheet2, ...).

Usage:
  python3 xlsx-helper.py book.xlsx  --to csv  --out ./out
  python3 xlsx-helper.py book.xlsx  --to md   --out book.md
  python3 xlsx-helper.py book.xlsx  --to md   > book.md
  python3 xlsx-helper.py book.xlsx  --to md   --sheet Summary
  python3 xlsx-helper.py table.csv  --to xlsx --out book.xlsx
  python3 xlsx-helper.py report.md  --to xlsx --out book.xlsx
"""
import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write(
        "error: openpyxl not installed.\n"
        "  pip install openpyxl       (or pip3 / pipx / your venv of choice)\n"
    )
    sys.exit(2)


def cell_to_str(value):
    if value is None:
        return ""
    # whole-number floats render as "3" not "3.0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def sheet_rows(ws):
    rows = [[cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    return rows


def slugify_sheet(name):
    bad = '<>:"/\\|?*'
    cleaned = "".join("_" if c in bad else c for c in name)
    cleaned = cleaned.strip().replace(" ", "_")
    return cleaned or "Sheet"


def md_escape(s):
    return s.replace("|", "\\|").replace("\n", " ")


def to_markdown(sheets):
    parts = []
    for name, rows in sheets:
        parts.append(f"## {name}")
        parts.append("")
        if not rows:
            parts.append("*(empty)*")
            parts.append("")
            continue
        ncols = max(len(r) for r in rows)
        rows = [r + [""] * (ncols - len(r)) for r in rows]
        header, body = rows[0], rows[1:]
        parts.append("| " + " | ".join(md_escape(c) for c in header) + " |")
        parts.append("|" + "|".join("---" for _ in range(ncols)) + "|")
        for r in body:
            parts.append("| " + " | ".join(md_escape(c) for c in r) + " |")
        parts.append("")
    return "\n".join(parts).rstrip() + "\n"


def to_json(sheets):
    payload = {"sheets": [{"name": n, "rows": r} for n, r in sheets]}
    return json.dumps(payload, ensure_ascii=False, indent=2) + "\n"


_SEP_CELL_RE = re.compile(r"^:?-+:?$")
_HEADING_RE = re.compile(r"^##\s+(.+?)\s*$")


def is_md_table_separator(line):
    s = line.strip()
    if "-" not in s:
        return False
    parts = [p.strip() for p in s.strip("|").split("|")]
    if not parts or any(not p for p in parts):
        return False
    return all(_SEP_CELL_RE.match(p) for p in parts)


def parse_md_row(line):
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    # trailing unescaped pipe is a table delimiter, not data
    if s.endswith("|") and not s.endswith("\\|"):
        s = s[:-1]
    cells = []
    cur = []
    i = 0
    while i < len(s):
        c = s[i]
        if c == "\\" and i + 1 < len(s) and s[i + 1] == "|":
            cur.append("|")
            i += 2
        elif c == "|":
            cells.append("".join(cur).strip())
            cur = []
            i += 1
        else:
            cur.append(c)
            i += 1
    cells.append("".join(cur).strip())
    return cells


def parse_md_tables(text):
    """Lenient pipe-table extractor.

    Returns [(sheet_name, rows), ...]. Sheet name is the most recent `##`
    heading above the table (consumed once); otherwise Sheet1, Sheet2, ...
    """
    lines = text.splitlines()
    sheets = []
    pending_heading = None
    anon = 0
    i = 0
    while i < len(lines):
        line = lines[i]
        h = _HEADING_RE.match(line)
        if h:
            pending_heading = h.group(1).strip()
            i += 1
            continue
        if (
            i + 1 < len(lines)
            and "|" in line
            and is_md_table_separator(lines[i + 1])
        ):
            header = parse_md_row(line)
            ncols = len(header)
            rows = [header]
            i += 2
            while i < len(lines) and lines[i].strip() and "|" in lines[i]:
                row = parse_md_row(lines[i])
                if len(row) < ncols:
                    row += [""] * (ncols - len(row))
                elif len(row) > ncols:
                    row = row[:ncols]
                rows.append(row)
                i += 1
            if pending_heading:
                base = pending_heading
                pending_heading = None
            else:
                anon += 1
                base = f"Sheet{anon}"
            sheets.append((base, rows))
            continue
        i += 1
    return sheets


def excel_safe_sheet_name(name, used):
    bad = '\\/?*[]:'
    cleaned = "".join("_" if c in bad else c for c in name).strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    n = 1
    while cleaned in used:
        n += 1
        suffix = f"_{n}"
        cleaned = base[: 31 - len(suffix)] + suffix
    used.add(cleaned)
    return cleaned


def csv_to_xlsx(src, out, delim):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = excel_safe_sheet_name(src.stem, set())
    with src.open(newline="", encoding="utf-8") as f:
        for row in csv.reader(f, delimiter=delim):
            ws.append(row)
    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def md_to_xlsx(src, out):
    from openpyxl import Workbook
    text = src.read_text(encoding="utf-8")
    sheets = parse_md_tables(text)
    if not sheets:
        sys.stderr.write(f"error: no pipe-tables found in {src}\n")
        sys.exit(1)
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for name, rows in sheets:
        ws = wb.create_sheet(title=excel_safe_sheet_name(name, used))
        for row in rows:
            ws.append(row)
    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_delimited(sheets, out_dir, stem, ext, delim):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    written = []
    multiple = len(sheets) > 1
    for name, rows in sheets:
        suffix = f"__{slugify_sheet(name)}" if multiple else ""
        path = out_dir / f"{stem}{suffix}.{ext}"
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter=delim)
            for r in rows:
                w.writerow(r)
        written.append(path)
    return written


def select_sheets(wb, sheet_arg):
    names = wb.sheetnames
    if sheet_arg is None:
        return names
    if sheet_arg.isdigit():
        idx = int(sheet_arg)
        if 0 <= idx < len(names):
            return [names[idx]]
    if sheet_arg in names:
        return [sheet_arg]
    sys.stderr.write(
        f"error: sheet '{sheet_arg}' not found. Available: {names}\n"
    )
    sys.exit(1)


def main():
    p = argparse.ArgumentParser(
        description="Convert between XLSX and CSV/TSV/Markdown/JSON."
    )
    p.add_argument("input", help="path to input file (.xlsx in read mode; .csv/.tsv/.md in write mode)")
    p.add_argument("--to", choices=["csv", "tsv", "md", "json", "xlsx"], required=True)
    p.add_argument(
        "--from",
        dest="from_",
        choices=["csv", "tsv", "md"],
        help="input format when --to xlsx (defaults to input file extension)",
    )
    p.add_argument(
        "--out",
        help="output path: directory for csv/tsv read mode; file for md/json/xlsx (omit for stdout in read mode)",
    )
    p.add_argument(
        "--sheet",
        help="restrict to a single sheet (read mode only; name or 0-based index)",
    )
    args = p.parse_args()

    src = Path(args.input)
    if not src.is_file():
        sys.stderr.write(f"error: file not found: {src}\n")
        return 1

    if args.to == "xlsx":
        if not args.out:
            sys.stderr.write("error: --out FILE is required for --to xlsx\n")
            return 1
        from_fmt = args.from_
        if not from_fmt:
            ext = src.suffix.lstrip(".").lower()
            from_fmt = {"csv": "csv", "tsv": "tsv", "md": "md", "markdown": "md"}.get(ext)
        if from_fmt not in ("csv", "tsv", "md"):
            sys.stderr.write(
                "error: --from must be csv/tsv/md when --to xlsx "
                f"(got {from_fmt!r}; could not infer from extension {src.suffix!r})\n"
            )
            return 1
        if from_fmt in ("csv", "tsv"):
            csv_to_xlsx(src, args.out, "," if from_fmt == "csv" else "\t")
        else:
            md_to_xlsx(src, args.out)
        sys.stderr.write(f"wrote {args.out}\n")
        return 0

    wb = load_workbook(src, read_only=True, data_only=True)
    sheet_names = select_sheets(wb, args.sheet)
    sheets = [(n, sheet_rows(wb[n])) for n in sheet_names]

    if args.to in ("csv", "tsv"):
        if not args.out:
            sys.stderr.write(f"error: --out DIR is required for --to {args.to}\n")
            return 1
        delim = "," if args.to == "csv" else "\t"
        for path in write_delimited(sheets, args.out, src.stem, args.to, delim):
            sys.stderr.write(f"wrote {path}\n")
        return 0

    text = to_markdown(sheets) if args.to == "md" else to_json(sheets)
    if args.out:
        Path(args.out).write_text(text, encoding="utf-8")
    else:
        sys.stdout.write(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
